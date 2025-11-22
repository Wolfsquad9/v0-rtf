import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("[v0] Starting Google Sheets template generation...")

# Create a new workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define color scheme (Dark Knight theme)
COLORS = {
    'primary': 'D4AF37',  # Gold
    'secondary': '8B9DC3',  # Blue-gray
    'dark_bg': '0A0A0A',  # Very dark
    'medium_bg': '1A1A1A',  # Dark gray
    'light_text': 'FFFFFF',  # White
    'accent': '00D4FF',  # Cyan
}

# Define styles
header_font = Font(name='Arial', size=14, bold=True, color=COLORS['light_text'])
subheader_font = Font(name='Arial', size=12, bold=True, color=COLORS['primary'])
normal_font = Font(name='Arial', size=10, color=COLORS['light_text'])
small_font = Font(name='Arial', size=9, color=COLORS['secondary'])

header_fill = PatternFill(start_color=COLORS['dark_bg'], end_color=COLORS['dark_bg'], fill_type='solid')
subheader_fill = PatternFill(start_color=COLORS['medium_bg'], end_color=COLORS['medium_bg'], fill_type='solid')
accent_fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)

print("[v0] Creating Overview sheet...")

# 1. OVERVIEW / COMMAND CENTER
overview = wb.create_sheet("📋 OVERVIEW")
overview.sheet_properties.tabColor = COLORS['primary']

overview['A1'] = "RETURN TO FORM - ULTIMATE EDITION"
overview['A1'].font = Font(name='Arial', size=20, bold=True, color=COLORS['primary'])
overview['A1'].alignment = center_align
overview.merge_cells('A1:H1')

overview['A3'] = "157+ Pages • 12 Weeks • Science-Backed Training"
overview['A3'].font = subheader_font
overview['A3'].alignment = center_align
overview.merge_cells('A3:H3')

overview['A5'] = "PRIMARY GOAL:"
overview['A5'].font = header_font
overview['B5'].font = normal_font
overview.merge_cells('B5:H5')

overview['A7'] = "SECONDARY METRIC (e.g., Weight, Lift PR):"
overview['A7'].font = header_font
overview['B7'] = "Starting Value:"
overview['C7'].font = normal_font
overview['D7'] = "Target Value:"
overview['E7'].font = normal_font

overview['A10'] = "NAVIGATION - QUICK LINKS"
overview['A10'].font = header_font
overview['A10'].fill = accent_fill
overview.merge_cells('A10:H10')

# Add navigation table
nav_headers = ['Section', 'Sheet Name', 'Description']
for col, header in enumerate(nav_headers, start=1):
    cell = overview.cell(row=11, column=col)
    cell.value = header
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = center_align

nav_items = [
    ('Core Metrics', '📊 METRICS', 'Starting measurements and progress photos'),
    ('RPE Reference', '🔥 RPE GUIDE', 'Rate of Perceived Exertion scale'),
    ('Week 1-12 Logs', 'Week 1-12 tabs', 'Weekly objectives and daily workout logs'),
    ('Final Summary', '🏆 SUMMARY', '12-week review and results'),
]

for row, (section, sheet, desc) in enumerate(nav_items, start=12):
    overview.cell(row=row, column=1).value = section
    overview.cell(row=row, column=2).value = sheet
    overview.cell(row=row, column=3).value = desc
    for col in range(1, 4):
        overview.cell(row=row, column=col).font = normal_font
        overview.cell(row=row, column=col).border = thin_border

# Set column widths
overview.column_dimensions['A'].width = 20
overview.column_dimensions['B'].width = 20
overview.column_dimensions['C'].width = 50
for col in ['D', 'E', 'F', 'G', 'H']:
    overview.column_dimensions[col].width = 15

print("[v0] Creating Core Metrics sheet...")

# 2. CORE METRICS
metrics = wb.create_sheet("📊 METRICS")
metrics.sheet_properties.tabColor = COLORS['accent']

metrics['A1'] = "CORE METRICS & MEASUREMENTS"
metrics['A1'].font = Font(name='Arial', size=16, bold=True, color=COLORS['primary'])
metrics.merge_cells('A1:D1')

# Measurement table
metrics['A3'] = "Measurement"
metrics['B3'] = "Start (Week 1)"
metrics['C3'] = "Final (Week 12)"
metrics['D3'] = "Change"

for col in ['A', 'B', 'C', 'D']:
    metrics[f'{col}3'].font = subheader_font
    metrics[f'{col}3'].fill = subheader_fill
    metrics[f'{col}3'].alignment = center_align

measurements = ['Weight (lbs/kg)', 'Waist (inches/cm)', 'Arms (inches/cm)', 
                'Chest (inches/cm)', 'Shoulders (inches/cm)', 'Body Fat %']

for row, measurement in enumerate(measurements, start=4):
    metrics[f'A{row}'] = measurement
    metrics[f'A{row}'].font = normal_font
    metrics[f'A{row}'].border = thin_border
    
    for col in ['B', 'C']:
        metrics[f'{col}{row}'].font = normal_font
        metrics[f'{col}{row}'].border = thin_border
        metrics[f'{col}{row}'].alignment = center_align
    
    # Formula for change
    metrics[f'D{row}'] = f'=C{row}-B{row}'
    metrics[f'D{row}'].font = Font(name='Arial', size=10, bold=True, color=COLORS['accent'])
    metrics[f'D{row}'].border = thin_border
    metrics[f'D{row}'].alignment = center_align

metrics['A11'] = "PROGRESS PHOTOS"
metrics['A11'].font = header_font
metrics['A11'].fill = accent_fill
metrics.merge_cells('A11:D11')

metrics['A12'] = "Start Photo (Front/Back/Side)"
metrics['A12'].font = normal_font
metrics.merge_cells('A12:B12')
metrics.row_dimensions[12].height = 100

metrics['C12'] = "End Photo (Front/Back/Side)"
metrics['C12'].font = normal_font
metrics.merge_cells('C12:D12')

metrics['A15'] = "VISION BOARD"
metrics['A15'].font = header_font
metrics['A15'].fill = accent_fill
metrics.merge_cells('A15:D15')

vision_items = ['Inspiration 1', 'Inspiration 2', 'Key Quote/Mantra', 
                'Reward for Completion', "What's at Stake?"]

for row, item in enumerate(vision_items, start=16):
    metrics[f'A{row}'] = item
    metrics[f'A{row}'].font = normal_font
    metrics.merge_cells(f'B{row}:D{row}')
    metrics[f'B{row}'].border = thin_border

metrics.column_dimensions['A'].width = 25
for col in ['B', 'C', 'D']:
    metrics.column_dimensions[col].width = 20

print("[v0] Creating RPE Guide sheet...")

# 3. RPE SCALE REFERENCE
rpe = wb.create_sheet("🔥 RPE GUIDE")
rpe.sheet_properties.tabColor = COLORS['secondary']

rpe['A1'] = "RPE SCALE REFERENCE"
rpe['A1'].font = Font(name='Arial', size=16, bold=True, color=COLORS['primary'])
rpe.merge_cells('A1:C1')

rpe['A3'] = "Rate of Perceived Exertion (RPE) helps track workout intensity and fatigue."
rpe['A3'].font = normal_font
rpe.merge_cells('A3:C3')

rpe_scale = [
    ('10', 'Failure / Max Effort', 'Cannot do another rep'),
    ('9', 'Very Hard', '1 Rep left in the tank'),
    ('8', 'Hard', '2 Reps left in the tank'),
    ('7', 'Moderate', 'Could do 3+ Reps'),
    ('6-', 'Easy', 'Warm-up, cardio, or skill work'),
]

rpe['A5'] = "RPE"
rpe['B5'] = "Intensity"
rpe['C5'] = "Description"

for col in ['A', 'B', 'C']:
    rpe[f'{col}5'].font = subheader_font
    rpe[f'{col}5'].fill = subheader_fill
    rpe[f'{col}5'].alignment = center_align

for row, (score, intensity, desc) in enumerate(rpe_scale, start=6):
    rpe[f'A{row}'] = score
    rpe[f'B{row}'] = intensity
    rpe[f'C{row}'] = desc
    
    for col in ['A', 'B', 'C']:
        rpe[f'{col}{row}'].font = normal_font
        rpe[f'{col}{row}'].border = thin_border
        rpe[f'{col}{row}'].alignment = left_align

rpe['A12'] = "DAILY RPE AUTO-CALCULATION"
rpe['A12'].font = header_font
rpe['A12'].fill = accent_fill
rpe.merge_cells('A12:C12')

rpe['A14'] = "Metric"
rpe['B14'] = "Score (1-10)"
rpe['C14'] = "Description"

for col in ['A', 'B', 'C']:
    rpe[f'{col}14'].font = subheader_font
    rpe[f'{col}14'].fill = subheader_fill
    rpe[f'{col}14'].alignment = center_align

daily_metrics = [
    ('Workout Intensity', 'How hard was your main session?'),
    ('Nutrition Adherence', 'How closely did you stick to your diet?'),
    ('Sleep Quality', 'Did you hit your sleep target?'),
    ('Hydration/Supplements', 'Water intake, vitamin consistency'),
    ('Mindset/Focus', 'Mental clarity, stress management'),
]

for row, (metric, desc) in enumerate(daily_metrics, start=15):
    rpe[f'A{row}'] = metric
    rpe[f'C{row}'] = desc
    
    for col in ['A', 'C']:
        rpe[f'{col}{row}'].font = normal_font
        rpe[f'{col}{row}'].border = thin_border
    
    rpe[f'B{row}'].border = thin_border
    rpe[f'B{row}'].alignment = center_align

rpe[f'A{row+1}'] = "TOTAL DAILY RPE"
rpe[f'A{row+1}'].font = Font(name='Arial', size=10, bold=True, color=COLORS['accent'])
rpe[f'B{row+1}'] = f'=SUM(B15:B{row})'
rpe[f'B{row+1}'].font = Font(name='Arial', size=10, bold=True, color=COLORS['accent'])
rpe[f'B{row+1}'].border = thin_border
rpe[f'B{row+1}'].alignment = center_align
rpe[f'C{row+1}'] = "Max 50"
rpe[f'C{row+1}'].font = small_font

rpe.column_dimensions['A'].width = 25
rpe.column_dimensions['B'].width = 15
rpe.column_dimensions['C'].width = 40

print("[v0] Creating weekly sheets (this may take a moment)...")

# 4. WEEKLY LOGS (12 weeks)
PLANNED_WEEKS = 12

for week in range(1, PLANNED_WEEKS + 1):
    print(f"[v0] Creating Week {week} sheet...")
    
    # Create weekly overview sheet
    ws = wb.create_sheet(f"Week {week}")
    ws.sheet_properties.tabColor = COLORS['primary']
    
    ws['A1'] = f"WEEK {week} LOG - THE CORE"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:H1')
    
    ws['A3'] = "PRIMARY FOCUS:"
    ws['A3'].font = header_font
    ws.merge_cells('B3:H3')
    ws['B3'].border = thin_border
    
    ws['A5'] = "DAILY HABITS/NON-NEGOTIABLES"
    ws['A5'].font = subheader_font
    ws['A5'].fill = subheader_fill
    ws.merge_cells('A5:H5')
    
    habits = ['Habit 1', 'Habit 2', 'Habit 3', 'Habit 4']
    for col, habit in enumerate(habits, start=1):
        ws.cell(row=6, column=col).value = habit
        ws.cell(row=6, column=col).font = normal_font
        ws.cell(row=7, column=col).value = "✓ or ✗"
        ws.cell(row=7, column=col).border = thin_border
        ws.cell(row=7, column=col).alignment = center_align
    
    ws['A9'] = "HABIT & RECOVERY TRACKER"
    ws['A9'].font = header_font
    ws['A9'].fill = accent_fill
    ws.merge_cells('A9:H9')
    
    # Habit tracker headers
    days = ['M', 'T', 'W', 'T', 'F', 'S', 'S']
    ws['A10'] = "HABIT / FOCUS"
    ws['A10'].font = subheader_font
    ws['A10'].fill = subheader_fill
    
    for col, day in enumerate(days, start=2):
        cell = ws.cell(row=10, column=col)
        cell.value = day
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
    
    # Habit rows
    habit_items = [
        'Sleep Goal Reached (7.5 hrs)',
        'Nutrition Compliant',
        'Hydration Target Met',
        'Mindset/Meditation/Reading',
    ]
    
    for row, habit in enumerate(habit_items, start=11):
        ws.cell(row=row, column=1).value = habit
        ws.cell(row=row, column=1).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        
        for col in range(2, 9):
            ws.cell(row=row, column=col).value = "☐"
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align
    
    ws[f'A{row+1}'] = "TOTAL WORKOUTS COMPLETED"
    ws[f'A{row+1}'].font = Font(name='Arial', size=10, bold=True, color=COLORS['accent'])
    ws.merge_cells(f'B{row+1}:H{row+1}')
    ws[f'B{row+1}'].value = "e.g., 5/7"
    ws[f'B{row+1}'].border = thin_border
    ws[f'B{row+1}'].alignment = center_align
    
    ws[f'A{row+3}'] = "WEEKLY REVIEW & STRATEGY"
    ws[f'A{row+3}'].font = header_font
    ws[f'A{row+3}'].fill = accent_fill
    ws.merge_cells(f'A{row+3}:H{row+3}')
    
    ws[f'A{row+4}'] = "What went right? What went wrong? What is the 1 adjustment for next week?"
    ws[f'A{row+4}'].font = normal_font
    ws.merge_cells(f'A{row+4}:H{row+7}')
    ws[f'A{row+4}'].border = thin_border
    ws[f'A{row+4}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row+4].height = 80
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    
    # Create daily workout sheets for each day
    days_full = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
    
    for day_idx, day_name in enumerate(days_full, start=1):
        ws_day = wb.create_sheet(f"W{week}D{day_idx}")
        ws_day.sheet_properties.tabColor = COLORS['secondary']
        
        ws_day['A1'] = f"{day_name}, WEEK {week}"
        ws_day['A1'].font = Font(name='Arial', size=14, bold=True, color=COLORS['primary'])
        ws_day.merge_cells('A1:F1')
        
        ws_day['A3'] = "Planned Training:"
        ws_day['A3'].font = subheader_font
        ws_day.merge_cells('B3:F3')
        ws_day['B3'].value = "e.g., Chest & Triceps / LISS Cardio"
        ws_day['B3'].border = thin_border
        
        ws_day['A5'] = "WORKOUT LOG"
        ws_day['A5'].font = header_font
        ws_day['A5'].fill = accent_fill
        ws_day.merge_cells('A5:F5')
        
        # Workout grid headers
        headers = ['EXERCISE', 'SET 1', 'SET 2', 'SET 3', 'SET 4', 'RPE / NOTES']
        for col, header in enumerate(headers, start=1):
            cell = ws_day.cell(row=6, column=col)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_align
        
        # 6 exercise rows
        for ex_row in range(7, 13):
            ws_day.cell(row=ex_row, column=1).value = f"Exercise {ex_row-6}"
            ws_day.cell(row=ex_row, column=1).font = small_font
            ws_day.cell(row=ex_row, column=1).border = thin_border
            
            for col in range(2, 7):
                ws_day.cell(row=ex_row, column=col).border = thin_border
                ws_day.cell(row=ex_row, column=col).alignment = center_align
        
        ws_day['A14'] = "DAILY RPE & NOTES"
        ws_day['A14'].font = header_font
        ws_day['A14'].fill = accent_fill
        ws_day.merge_cells('A14:F14')
        
        ws_day['A16'] = "Daily Performance Score (1-10)"
        ws_day['A16'].font = subheader_font
        ws_day.merge_cells('A16:F16')
        
        # Daily RPE metrics
        rpe_metrics = [
            'Workout Intensity',
            'Nutrition Adherence',
            'Sleep Quality',
            'Hydration/Supplements',
            'Mindset/Focus',
        ]
        
        for row, metric in enumerate(rpe_metrics, start=17):
            ws_day.cell(row=row, column=1).value = metric
            ws_day.cell(row=row, column=1).font = normal_font
            ws_day.cell(row=row, column=1).border = thin_border
            
            ws_day.cell(row=row, column=2).value = "/10"
            ws_day.cell(row=row, column=2).border = thin_border
            ws_day.cell(row=row, column=2).alignment = center_align
        
        ws_day[f'A{row+1}'] = "TOTAL DAILY RPE (Max 50)"
        ws_day[f'A{row+1}'].font = Font(name='Arial', size=10, bold=True, color=COLORS['accent'])
        ws_day[f'B{row+1}'] = f'=SUM(B17:B{row})'
        ws_day[f'B{row+1}'].font = Font(name='Arial', size=10, bold=True, color=COLORS['accent'])
        ws_day[f'B{row+1}'].border = thin_border
        ws_day[f'B{row+1}'].alignment = center_align
        
        ws_day[f'A{row+3}'] = "Daily Notes & Lessons Learned"
        ws_day[f'A{row+3}'].font = subheader_font
        ws_day.merge_cells(f'A{row+3}:F{row+6}')
        ws_day[f'A{row+3}'].border = thin_border
        ws_day[f'A{row+3}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws_day.row_dimensions[row+3].height = 80
        
        # Set column widths
        ws_day.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_day.column_dimensions[col].width = 15

print("[v0] Creating Final Summary sheet...")

# 5. FINAL SUMMARY
summary = wb.create_sheet("🏆 SUMMARY")
summary.sheet_properties.tabColor = COLORS['accent']

summary['A1'] = "12-WEEK SUMMARY - MISSION COMPLETE"
summary['A1'].font = Font(name='Arial', size=16, bold=True, color=COLORS['primary'])
summary.merge_cells('A1:D1')

summary['A3'] = "FINAL REFLECTION & KEY TAKEAWAYS"
summary['A3'].font = header_font
summary['A3'].fill = accent_fill
summary.merge_cells('A3:D3')

summary['A4'] = "Summarize your 12 weeks: What was your single greatest win? What was the hardest lesson?"
summary['A4'].font = normal_font
summary.merge_cells('A4:D8')
summary['A4'].border = thin_border
summary['A4'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
summary.row_dimensions[4].height = 100

summary['A10'] = "QUANTIFIABLE RESULTS"
summary['A10'].font = header_font
summary['A10'].fill = accent_fill
summary.merge_cells('A10:D10')

summary['A11'] = "Metric"
summary['B11'] = "Start"
summary['C11'] = "End"
summary['D11'] = "Change"

for col in ['A', 'B', 'C', 'D']:
    summary[f'{col}11'].font = subheader_font
    summary[f'{col}11'].fill = subheader_fill
    summary[f'{col}11'].alignment = center_align

final_metrics = ['Weight', 'Key Lift PR (e.g., Bench)', 'Total Workouts Completed']

for row, metric in enumerate(final_metrics, start=12):
    summary[f'A{row}'] = metric
    summary[f'A{row}'].font = normal_font
    summary[f'A{row}'].border = thin_border
    
    for col in ['B', 'C']:
        summary[f'{col}{row}'].border = thin_border
        summary[f'{col}{row}'].alignment = center_align
    
    summary[f'D{row}'] = f'=C{row}-B{row}'
    summary[f'D{row}'].font = Font(name='Arial', size=10, bold=True, color=COLORS['accent'])
    summary[f'D{row}'].border = thin_border
    summary[f'D{row}'].alignment = center_align

summary['A16'] = "NEXT MISSION / FUTURE GOALS"
summary['A16'].font = header_font
summary['A16'].fill = accent_fill
summary.merge_cells('A16:D16')

summary['A17'] = "Where do you go from here? What is the next major challenge?"
summary['A17'].font = normal_font
summary.merge_cells('A17:D20')
summary['A17'].border = thin_border
summary['A17'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
summary.row_dimensions[17].height = 80

summary.column_dimensions['A'].width = 30
for col in ['B', 'C', 'D']:
    summary.column_dimensions[col].width = 20

# Save the workbook
filename = f"Return_to_Form_Ultimate_Edition_{PLANNED_WEEKS}Weeks_GoogleSheets.xlsx"
wb.save(filename)

print(f"[v0] ✅ Google Sheets template created successfully!")
print(f"[v0] 📁 File saved as: {filename}")
print(f"[v0] 📊 Total sheets created: {len(wb.sheetnames)}")
print(f"[v0] 🎯 Includes: Overview, Metrics, RPE Guide, {PLANNED_WEEKS} weekly logs with daily workout sheets, and Final Summary")
print(f"[v0] 💡 To use: Upload this Excel file to Google Drive and open with Google Sheets")
print(f"[v0] 🛍️ Ready for Etsy: Professional formatting, auto-calculations, and comprehensive tracking")
